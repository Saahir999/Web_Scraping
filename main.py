from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import pandas as pd
import openpyxl

months = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
year_start = 2007
year_end = 2023
FILENAME = "comp.xlsx"
# Do not change -- acc to site reqs
countries = ["SAUDI ARAB", "U ARAB EMTS", "IRAQ", "OMAN", "KUWAIT", "QATAR",
             "TURKEY", "UKRAINE", "VENEZUELA", "YEMEN REPUBLIC", "LIBERIA",
             "U S A", "SYRIA", "BRAZIL", "EGYPT A RP", "MALAYSIA", "COLOMBIA",
             "NIGERIA", "RUSSIA", "IRAN", "MEXICO", "SUDAN", "Total"]
country_DICT = {}
for c in countries:
    country_DICT[c] = []
# country_DICT = {"SAUDI ARAB": [], "U ARAB EMTS": [], "IRAQ": [], "OMAN": [], "KUWAIT": [], "QATAR": [], "NIGERIA": [], "Total": []}
table_row_start = 2
table_row_end = 50

table_col_start = 2
table_col_end = 8

# Start the Selenium webdriver and init workbook
driver = webdriver.Chrome()
# writer = pd.ExcelWriter('Data2.xlsx', engine='xlsxwriter')
# workbook = writer.book

# Open the website
driver.get('https://tradestat.commerce.gov.in/meidb/comcntq.asp?ie=i')
# writer = pd.ExcelWriter(FILENAME, mode="a", engine="openpyxl", if_sheet_exists="overlay")
workbook = openpyxl.load_workbook(FILENAME)


def table_path():
    return str('//table/tbody/tr[' + str(i) + ']/td[' + str(j) + ']')


final_df = None
try:
    # if str() not in workbook.sheetnames:
    #     workbook.create_sheet(title=country)
    #     workbook.save(FILENAME)
    # sheet = workbook[country]
    for year in range(year_start, year_end + 1):
        for month in months:
            # sheet.append(["Date",
            #               "Country",
            #               str(month).lower() + " " + str(year - 1) + " (R)",
            #               str(month).lower() + " " + str(year) + " (R)",
            #               "%YearGrowth",
            #               "Apr-" + str(month).lower() + " " + str(year - 1) + " (R)",
            #               "Apr-" + str(month).lower() + " " + str(year) + " (R)",
            #               "%FinYearGrowth"])
            # workbook.save(FILENAME)

            # Choose Month
            dropdown_month = Select(driver.find_element(By.NAME, 'Mm1'))
            dropdown_month.select_by_visible_text(month)

            # Choose year
            dropdown_month = Select(driver.find_element(By.NAME, 'yy1'))
            dropdown_month.select_by_visible_text(str(year))

            # Enter oil code
            driver.find_element(By.NAME, 'hscode').clear()
            enter = driver.find_element(By.NAME, 'hscode')
            enter.send_keys("2709")

            # Button Press
            radio = driver.find_element(By.NAME, 'radiousd')
            radio.click()

            # # Button Press
            # radio = driver.find_element(By.NAME, 'radioqty')
            # radio.click()

            # Button Press
            button = driver.find_element(By.NAME, 'button1')
            button.click()

            # Wait for the page to load
            driver.implicitly_wait(5)

            # Extract the numbers from the resulting page
            row_data = []
            month_df = []
            complete = 0

            for i in range(table_row_start, table_row_end + 1):
                row_data = []
                c = ""
                try:
                    for j in range(table_col_start, table_col_end + 1):
                        row_elements = driver.find_element(By.XPATH, table_path()).text
                        if row_elements not in countries and j == table_col_start:
                            break
                        # print(row_elements)
                        if j == table_col_start:
                            c = row_elements
                            complete += 1
                        row_data.append(row_elements)
                except Exception as e:
                    print(str(i) + " row xpath not found")
                    if i > 1:
                        break

                if row_data:
                    # print(row_data)
                    # append row to excel
                    df = [str(month).lower() + " " + str(year)] + row_data
                    try:
                        country_DICT[c].append(df)
                        # sheet.append(df)
                        # workbook.save(FILENAME)
                        # print(df)
                        # month_df = pd.concat([df, month_df])
                        month_df.append(df)
                    except Exception as e:
                        print("APPEND ERROR")
                        print("month data is ")
                        print(type(month_df))
                        print(month_df)

                        print(e)

                if complete == len(countries):
                    break

                # count row length of month_df and store for future offset, enter it into year sheet

            if len(month_df):
                print(month_df)
                # print("https://wrote_data")

            driver.back()

        # sheet.append([''] * 10)
        # workbook.save(FILENAME)
        # final_df.to_excel(writer, sheet_name=str(year), startrow=writer.sheets[str(year)].max_row)

    for nation, data in country_DICT.items():
        print(data)
        if nation not in workbook.sheetnames:
            workbook.create_sheet(title=nation)
            workbook.save(FILENAME)
        sheet = workbook[nation]
        for d in data:
            sheet.append(d)
            workbook.save(FILENAME)

except Exception as e:
    for nation, data in country_DICT.items():
        print(data)
        if nation not in workbook.sheetnames:
            workbook.create_sheet(title=nation)
            workbook.save(FILENAME)
        sheet = workbook[nation]
        for d in data:
            sheet.append(d)
            workbook.save(FILENAME)
    print("Scrape aborted --Saving")
    print(e)

# Close the browser
driver.quit()
